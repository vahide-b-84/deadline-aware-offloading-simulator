# post_process_results.py (DEADLINE-BASED) - FINAL (TaskDist sheets include 4 metrics)
# - Enriches each result .xlsx by adding:
#     * "TaskDist_LastEp" sheet + 4 charts (Load%, OnTime_Rate, Deadline_Miss_Rate, Avg_Tardiness)
#     * "TaskDist_Last40" sheet + 4 charts (Load%, OnTime_Rate, Deadline_Miss_Rate, Avg_Tardiness)
# - Creates ONE root-level Final_Result_All.xlsx
#     * one sheet per results folder
#     * compares:
#         Avg Reward
#         AVG_OnTime_Rate
#         AVG_Deadline_Miss_Rate
#         AVG_Tardiness
#       across models (line charts)

import os
import re
import zipfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference

from config.paths import RESULTS_DIR, ensure_dirs


# ----------------------------
# Helpers
# ----------------------------

def is_valid_excel(file_path: str) -> bool:
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            return "[Content_Types].xml" in archive.namelist()
    except zipfile.BadZipFile:
        return False


def safe_sheet_delete(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]


def write_df_to_sheet(ws, df: pd.DataFrame, start_row=1, start_col=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)


def _header_to_col(ws) -> dict:
    """Return mapping {header(str): 1-based col index} from first row."""
    out = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        out[str(cell.value).strip()] = cell.col_idx
    return out


def _add_server_label_column(ws, server_ids, server_types, label_col: int):
    ws.cell(row=1, column=label_col, value="Server_Label")
    for i, (sid, stype) in enumerate(zip(server_ids, server_types), start=2):
        ws.cell(row=i, column=label_col, value=f"{sid} ({stype})")


def add_taskdist_4charts(ws, df: pd.DataFrame, title_suffix: str,
                         anchor1="J1", anchor2="J20", anchor3="J39", anchor4="J58"):
    """
    Adds 4 bar charts on TaskDist sheets:
      1) Load%                 (Primary_Tasks_Percentage)
      2) OnTime_Rate           (OnTime_Rate)
      3) Deadline_Miss_Rate    (Deadline_Miss_Rate)
      4) Avg_Tardiness         (Avg_Tardiness)

    Expects df columns:
      Server_ID, Server_Type,
      Primary_Tasks_Percentage, Primary_Task_Count,
      OnTime_Count, Deadline_Miss_Count,
      OnTime_Rate, Deadline_Miss_Rate,
      Avg_Delay, Avg_Deadline, Avg_Tardiness
    """
    n = len(df)
    if n <= 0:
        return

    server_ids = df["Server_ID"].tolist()
    server_types = df["Server_Type"].tolist()

    # Put label column far enough to the right to avoid clashing with df
    label_col = max(1, ws.max_column + 2)
    _add_server_label_column(ws, server_ids, server_types, label_col=label_col)
    categories = Reference(ws, min_col=label_col, min_row=2, max_row=n + 1)

    colmap = _header_to_col(ws)

    def _ref(col_name: str):
        if col_name not in colmap:
            return None
        c = colmap[col_name]
        return Reference(ws, min_col=c, max_col=c, min_row=1, max_row=n + 1)

    specs = [
        ("Primary_Tasks_Percentage", f"Load% ({title_suffix})", "Percent", anchor1, 10),
        ("OnTime_Rate", f"OnTimeRate ({title_suffix})", "OnTime_Rate", anchor2, 11),
        ("Deadline_Miss_Rate", f"DeadlineMissRate ({title_suffix})", "Deadline_Miss_Rate", anchor3, 12),
        ("Avg_Tardiness", f"AvgTardiness ({title_suffix})", "Avg_Tardiness", anchor4, 13),
    ]

    for col_name, title, ytitle, anchor, style in specs:
        data_ref = _ref(col_name)
        if data_ref is None:
            continue
        ch = BarChart()
        ch.type = "col"
        ch.style = style
        ch.title = title
        ch.y_axis.title = ytitle
        ch.x_axis.title = "Server (Type)"
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(categories)
        ch.width = 22
        ch.height = 10
        ws.add_chart(ch, anchor)


def read_required_sheets(xlsx_path: str):
    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    servers_df = pd.read_excel(xls, sheet_name="Servers")
    tasks_df = pd.read_excel(xls, sheet_name="TaskAssignments")
    return servers_df, tasks_df


def _pick_status_col(tasks_df: pd.DataFrame) -> str:
    cols = [str(c).strip() for c in tasks_df.columns]
    if "Final_status" in cols:
        return "Final_status"
    return "Primary_Status"


def _ensure_delay_and_deadline(tasks_df: pd.DataFrame) -> pd.DataFrame:
    df = tasks_df.copy()
    cols = [str(c).strip() for c in df.columns]
    if "Delay" not in cols:
        if "Primary_Start" in cols and "Primary_End" in cols:
            df["Delay"] = df["Primary_End"] - df["Primary_Start"]
        else:
            df["Delay"] = float("nan")
    if "Deadline" not in cols:
        df["Deadline"] = float("nan")
    return df


def compute_server_task_distribution_and_rates(servers_df: pd.DataFrame, tasks_df: pd.DataFrame, episode_ids):
    """
    Per-server metrics for given episode(s):
      Primary_Tasks_Percentage, Primary_Task_Count,
      OnTime_Count, Deadline_Miss_Count,
      OnTime_Rate, Deadline_Miss_Rate
    """
    server_ids = sorted(servers_df["Server_ID"].astype(int).tolist())
    server_types = servers_df.set_index("Server_ID").loc[server_ids]["Server_Type"]

    dfw = tasks_df[tasks_df["episode"].isin(list(episode_ids))].copy()
    status_col = _pick_status_col(dfw) if not dfw.empty else "Primary_Status"

    primary_count = {sid: 0 for sid in server_ids}
    ontime_count = {sid: 0 for sid in server_ids}
    miss_count = {sid: 0 for sid in server_ids}

    if not dfw.empty:
        for _, row in dfw.iterrows():
            if pd.isna(row.get("Primary")):
                continue
            sid = int(row["Primary"])
            if sid not in primary_count:
                continue

            primary_count[sid] += 1
            status = str(row.get(status_col, "")).strip().lower()
            if status == "success":
                ontime_count[sid] += 1
            elif status == "failure":
                miss_count[sid] += 1

    total_tasks = sum(primary_count.values())
    primary_pct = [(primary_count[sid] / total_tasks) * 100 for sid in server_ids] if total_tasks > 0 else [0 for _ in server_ids]

    ontime_rate = []
    miss_rate = []
    for sid in server_ids:
        denom = primary_count[sid]
        ontime_rate.append((ontime_count[sid] / denom) if denom > 0 else 0.0)
        miss_rate.append((miss_count[sid] / denom) if denom > 0 else 0.0)

    return pd.DataFrame({
        "Server_ID": server_ids,
        "Server_Type": [server_types[sid] for sid in server_ids],
        "Primary_Tasks_Percentage": primary_pct,
        "Primary_Task_Count": [primary_count[sid] for sid in server_ids],
        "OnTime_Count": [ontime_count[sid] for sid in server_ids],
        "Deadline_Miss_Count": [miss_count[sid] for sid in server_ids],
        "OnTime_Rate": ontime_rate,
        "Deadline_Miss_Rate": miss_rate,
    })


def compute_server_quality_metrics(servers_df: pd.DataFrame, tasks_df: pd.DataFrame, episode_ids):
    """
    Per-server quality metrics over given episode(s):
      Avg_Delay, Avg_Deadline, Avg_Tardiness
    where tardiness = max(0, Delay - Deadline)
    """
    server_ids = sorted(servers_df["Server_ID"].astype(int).tolist())
    server_types = servers_df.set_index("Server_ID").loc[server_ids]["Server_Type"]

    dfw = tasks_df[tasks_df["episode"].isin(list(episode_ids))].copy()
    dfw = _ensure_delay_and_deadline(dfw)

    task_count = {sid: 0 for sid in server_ids}
    sum_delay = {sid: 0.0 for sid in server_ids}
    sum_deadline = {sid: 0.0 for sid in server_ids}
    sum_tard = {sid: 0.0 for sid in server_ids}

    if not dfw.empty:
        for _, row in dfw.iterrows():
            if pd.isna(row.get("Primary")):
                continue
            sid = int(row["Primary"])
            if sid not in task_count:
                continue

            task_count[sid] += 1

            delay = row.get("Delay")
            deadline = row.get("Deadline")

            try:
                delay_val = float(delay)
            except Exception:
                delay_val = float("nan")

            try:
                deadline_val = float(deadline)
            except Exception:
                deadline_val = float("nan")

            if pd.notna(delay_val):
                sum_delay[sid] += delay_val
            if pd.notna(deadline_val):
                sum_deadline[sid] += deadline_val
            if pd.notna(delay_val) and pd.notna(deadline_val):
                sum_tard[sid] += max(0.0, delay_val - deadline_val)

    rows = []
    for sid in server_ids:
        n = task_count[sid]
        avg_delay = (sum_delay[sid] / n) if n > 0 else 0.0
        avg_deadline = (sum_deadline[sid] / n) if n > 0 else 0.0
        avg_tard = (sum_tard[sid] / n) if n > 0 else 0.0

        rows.append({
            "Server_ID": sid,
            "Server_Type": str(server_types[sid]),
            "Avg_Delay": avg_delay,
            "Avg_Deadline": avg_deadline,
            "Avg_Tardiness": avg_tard,
        })

    return pd.DataFrame(rows)


def process_one_result_file(xlsx_path: str, window_last_n: int = 40):
    if not is_valid_excel(xlsx_path):
        raise ValueError("Not a valid Excel file.")

    servers_df, tasks_df = read_required_sheets(xlsx_path)

    if tasks_df.empty or "episode" not in tasks_df.columns:
        raise ValueError("TaskAssignments is empty or missing 'episode' column.")

    episodes = sorted(pd.Series(tasks_df["episode"]).dropna().astype(int).unique().tolist())
    if not episodes:
        raise ValueError("No valid episode values found in TaskAssignments.")

    last_episode = episodes[-1]
    last_window = episodes[-min(window_last_n, len(episodes)):]
    window_label = f"{last_window[0]}..{last_window[-1]} (n={len(last_window)})"

    # --- TaskDist (Last Episode) ---
    df_dist_last = compute_server_task_distribution_and_rates(servers_df, tasks_df, [last_episode])
    df_q_last = compute_server_quality_metrics(servers_df, tasks_df, [last_episode])
    df_last = df_dist_last.merge(df_q_last, on=["Server_ID", "Server_Type"], how="left")

    # --- TaskDist (Last 40) ---
    df_dist_lastN = compute_server_task_distribution_and_rates(servers_df, tasks_df, last_window)
    df_q_lastN = compute_server_quality_metrics(servers_df, tasks_df, last_window)
    df_lastN = df_dist_lastN.merge(df_q_lastN, on=["Server_ID", "Server_Type"], how="left")

    wb = load_workbook(xlsx_path)

    # IMPORTANT: short names <= 31 chars
    S1 = "TaskDist_LastEp"
    S2 = "TaskDist_Last40"

    safe_sheet_delete(wb, S1)
    ws1 = wb.create_sheet(S1)
    write_df_to_sheet(ws1, df_last)
    add_taskdist_4charts(ws1, df_last, title_suffix=f"Ep={last_episode}",
                         anchor1="O2", anchor2="O21", anchor3="O40", anchor4="O59")

    safe_sheet_delete(wb, S2)
    ws2 = wb.create_sheet(S2)
    write_df_to_sheet(ws2, df_lastN)
    add_taskdist_4charts(ws2, df_lastN, title_suffix=f"Win {window_label}",
                         anchor1="O2", anchor2="O21", anchor3="O40", anchor4="O59")

    # Cleanup legacy names (optional)
    safe_sheet_delete(wb, "Recovery Strategy Distribution")
    safe_sheet_delete(wb, "Task Distribution")
    safe_sheet_delete(wb, "Task Distribution (Last Episode)")
    safe_sheet_delete(wb, "Task Distribution (Last 40 Episodes)")
    safe_sheet_delete(wb, "Server Quality (Rolling 40)")
    safe_sheet_delete(wb, "SrvQuality_R40")  # if created by older versions

    wb.save(xlsx_path)
    return True
# ----------------------------
# Final_Result_All builder
# ----------------------------

def model_label_from_filename(filename: str) -> str:
    base = os.path.splitext(filename)[0]
    parts = base.split("_")
    return parts[0].strip().lower() if parts else base.strip().lower()


def extract_logs_and_summary_metrics(xlsx_path: str):
    """
    Reads:
      Logs: Episode, Avg Reward
      Summary: Episode/episode,
        AVG_Deadline_Miss_Rate

    Returns:
      logs_df: Episode + AvgReward_<model> (rename done later)
      met_df : Episode + metrics (rename done later)
    """
    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")

    logs_df = pd.read_excel(xls, sheet_name="Logs")
    logs_df.columns = [str(c).strip() for c in logs_df.columns]
    if "Episode" not in logs_df.columns and "episode" in logs_df.columns:
        logs_df = logs_df.rename(columns={"episode": "Episode"})
    if "Episode" not in logs_df.columns:
        raise ValueError("Logs missing Episode column")
    if "Avg Reward" not in logs_df.columns:
        raise ValueError("Logs missing 'Avg Reward' column")
    logs_df = logs_df[["Episode", "Avg Reward"]].copy()

    sum_df = pd.read_excel(xls, sheet_name="Summary")
    sum_df.columns = [str(c).strip() for c in sum_df.columns]
    if "Episode" not in sum_df.columns and "episode" in sum_df.columns:
        sum_df = sum_df.rename(columns={"episode": "Episode"})
    if "Episode" not in sum_df.columns:
        raise ValueError("Summary missing Episode/episode column")

    # Ensure columns exist; if missing, create NaN to avoid crash
    need_cols = ["AVG_Deadline_Miss_Rate"]
    met_df = sum_df[["Episode"]].copy()
    for c in need_cols:
        if c in sum_df.columns:
            met_df[c] = sum_df[c]
        else:
            met_df[c] = float("nan")

    return logs_df, met_df


def _add_line_chart_cols(ws, title, y_title, episode_col, series_cols, max_row, anchor):
    """
    Adds one line chart with x=Episode and multiple y series (one column per series).
    Each series column must include a header in row 1.
    """
    if max_row < 2 or not series_cols:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Episode"

    cats = Reference(ws, min_col=episode_col, min_row=2, max_row=max_row)

    for c in series_cols:
        data = Reference(ws, min_col=c, max_col=c, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)

    chart.set_categories(cats)
    chart.width = 26
    chart.height = 12
    ws.add_chart(chart, anchor)

def build_final_result_all(root_dir: str, folder_payloads: dict):
    """
    Creates Final_Result_All.xlsx in root_dir with one sheet per results folder.

    Each sheet contains ONE table:
      Episode |
      AvgReward_<model>... |
      AVG_Deadline_Miss_Rate_<model>...

    Charts (no overlap):
      - Reward chart at H2
      - Miss chart at H30
    """
    if not folder_payloads:
        print("[WARN] No folders to summarize.")
        return

    out_path = os.path.join(root_dir, "Final_Result_All.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    for folder_name, payload in folder_payloads.items():
        sheet_name = folder_name[:31]
        ws = wb.create_sheet(sheet_name)

        merged = payload.get("merged")
        if merged is None or merged.empty:
            ws["A1"] = "No data"
            continue

        merged = merged.sort_values("Episode")
        write_df_to_sheet(ws, merged, start_row=1, start_col=1)

        max_row = 1 + merged.shape[0]
        header = list(merged.columns)

        def col_index(name: str):
            return header.index(name) + 1  # 1-based

        if "Episode" not in header:
            ws["A1"] = "Missing Episode column"
            continue

        c_episode = col_index("Episode")

        
        # --- Chart 1: Avg Reward (ONLY AvgReward_<model> columns)
        reward_cols = [c for c in header if str(c).lower().startswith("avgreward_")]
        if reward_cols:
            series_cols = [col_index(c) for c in reward_cols]
            _add_line_chart_cols(
                ws,
                title="Avg Reward Over Episodes (Compare Models)",
                y_title="Avg Reward",
                episode_col=c_episode,
                series_cols=series_cols,
                max_row=max_row,
                anchor="H2",
            )

        # --- Chart 2: AVG_Deadline_Miss_Rate (ONLY AVG_Deadline_Miss_Rate_<model> columns)
        miss_cols = [c for c in header if str(c).lower().startswith("avg_deadline_miss_rate_")]
        if miss_cols:
            series_cols = [col_index(c) for c in miss_cols]
            _add_line_chart_cols(
                ws,
                title="AVG_Deadline_Miss_Rate Over Episodes (Compare Models)",
                y_title="AVG_Deadline_Miss_Rate",
                episode_col=c_episode,
                series_cols=series_cols,
                max_row=max_row,
                anchor="H30",
            )

    wb.save(out_path)
    print(f"[OK] Created root summary: {out_path}")


# ----------------------------
# Folder scanning
# ----------------------------

def is_results_folder(name: str) -> bool:
    return bool(re.fullmatch(r"(homogeneous|heterogeneous)_results", name))


def is_result_xlsx(name: str) -> bool:
    return (
        name.lower().endswith(".xlsx")
        and not name.startswith("~$")
        and name.lower() != "final_result.xlsx"
        and name.lower() != "final_result_all.xlsx"
    )


def process_all_results(root_dir: str):
    folder_payloads = {}

    for name in os.listdir(root_dir):
        folder_path = os.path.join(root_dir, name)
        if not os.path.isdir(folder_path):
            continue
        if not is_results_folder(name):
            continue

        print(f"\n=== Processing folder: {name} ===")

        xlsx_files = [fn for fn in os.listdir(folder_path) if is_result_xlsx(fn)]
        if not xlsx_files:
            print("  [WARN] No xlsx files found.")
            continue

        # 1) Enrich each result xlsx
        for fn in xlsx_files:
            xlsx_path = os.path.join(folder_path, fn)
            try:
                process_one_result_file(xlsx_path, window_last_n=40)
                print(f"  [OK] enriched: {fn}")
            except Exception as e:
                print(f"  [FAIL] {fn} -> {e}")

        # 2) Merge metrics across models for this folder
        merged_all = None

        for fn in xlsx_files:
            xlsx_path = os.path.join(folder_path, fn)
            model = model_label_from_filename(fn)

            try:
                logs_df, met_df = extract_logs_and_summary_metrics(xlsx_path)
            except Exception as e:
                print(f"  [WARN] Skip in summary merge: {fn} -> {e}")
                continue

            logs_df = logs_df.rename(columns={"Avg Reward": f"AvgReward_{model}"})
            met_df = met_df.rename(columns={
                "AVG_Deadline_Miss_Rate": f"AVG_Deadline_Miss_Rate_{model}",
            })

            df_model = pd.merge(logs_df, met_df, on="Episode", how="outer")

            merged_all = df_model if merged_all is None else pd.merge(merged_all, df_model, on="Episode", how="outer")

        if merged_all is not None:
            merged_all = merged_all.sort_values("Episode")

        folder_payloads[name] = {"merged": merged_all}

    # 3) Write ONE root summary file
    build_final_result_all(root_dir, folder_payloads)


def main():
    ensure_dirs()
    root = RESULTS_DIR
    process_all_results(root)
    print("\nDONE.")


if __name__ == "__main__":
    main()
