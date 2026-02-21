# save_parameters_and_logs.py
# - deadline-based (NO failure model)
# - Reads Excel input files ONLY from data/
# - Writes results per (scenario, model)
# - Creates Excel-native charts (no PNG files)
# - Adds Summary.OnTime_Rate / AVG_OnTime_Rate (rolling 40)
# - Adds Summary.Deadline_Miss_Rate / AVG_Deadline_Miss_Rate (rolling 40)

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

from config.paths import DATA_DIR, RESULTS_DIR, ensure_dirs


def save_params_and_logs(params, log_data, task_Assignments_info):
    ensure_dirs()

    scenario = getattr(params, "SCENARIO_TYPE", "heterogeneous")
    model_name = str(getattr(params, "model_summary", "model")).strip().lower()

    # ---------------------------
    # Results folder + filename
    # ---------------------------
    results_dir = os.path.join(RESULTS_DIR, f"{scenario}_results")
    os.makedirs(results_dir, exist_ok=True)

    filename = os.path.join(results_dir, f"{model_name}_{scenario}.xlsx")

    # ---------------------------
    # Load Servers (from data/)
    # ---------------------------
    servers_xlsx_name = "homogeneous_server_info.xlsx" if scenario == "homogeneous" else "heterogeneous_server_info.xlsx"
    servers_path = os.path.join(DATA_DIR, servers_xlsx_name)
    if not os.path.exists(servers_path):
        raise FileNotFoundError(f"File not found: {servers_path}")

    # NEW: single sheet
    try:
        server_info = pd.read_excel(servers_path, sheet_name="Servers")
    except ValueError as e:
        raise ValueError(
            f"Sheet 'Servers' not found in '{servers_path}'. "
            f"Check your generator output sheet names."
        ) from e

    # ---------------------------
    # Load Tasks (from data/)
    # ---------------------------
    task_path = os.path.join(DATA_DIR, "task_parameters.xlsx")
    if not os.path.exists(task_path):
        raise FileNotFoundError(f"File not found: {task_path}")
    task_df = pd.read_excel(task_path)

    # ---------------------------
    # Params dataframe
    # ---------------------------
    params_data = {attr: [value] for attr, value in vars(params).items()}
    df_params = pd.DataFrame(params_data).transpose().reset_index()
    df_params.columns = ["Parameter", "Value"]

    # ---------------------------
    # Logs dataframe
    # expected: (episode, avg_reward, episodic_reward, avg_delay)
    # ---------------------------
    logs_rows = []
    for log in log_data:
        logs_rows.append({
            "Episode": log[0],
            "Avg Reward": log[1] if len(log) > 1 else None,
            "Episode Reward": log[2] if len(log) > 2 else None,
            "Avg Delay": log[3] if len(log) > 3 else None,
        })
    df_logs = pd.DataFrame(logs_rows)

    # ---------------------------
    # TaskAssignments dataframe (deadline-based)
    # expected tuple:
    # (episode, task_id, primary_id, primary_start, primary_end, primary_status, deadline)
    # ---------------------------
    df_task_Assignments = pd.DataFrame(
        task_Assignments_info,
        columns=[
            "episode",
            "task_id",
            "Primary",
            "Primary_Start",
            "Primary_End",
            "Primary_Status",
            "Deadline",
        ],
    )

    # add Delay + Final_status
    if not df_task_Assignments.empty:
        df_task_Assignments["Delay"] = df_task_Assignments["Primary_End"] - df_task_Assignments["Primary_Start"]
        # final status equals primary status (single server)
        df_task_Assignments["Final_status"] = df_task_Assignments["Primary_Status"]
    else:
        df_task_Assignments["Delay"] = []
        df_task_Assignments["Final_status"] = []

    # ---------------------------
    # Summary per episode + rolling rates (40)
    # ---------------------------
    if not df_task_Assignments.empty:
        summary_df = df_task_Assignments.groupby(["episode", "Final_status"]).size().unstack(fill_value=0)
    else:
        summary_df = pd.DataFrame()

    # Normalize expected columns
    if "success" not in summary_df.columns:
        summary_df["success"] = 0
    if "failure" not in summary_df.columns:
        summary_df["failure"] = 0

    summary_df = summary_df.rename(columns={"success": "OnTime", "failure": "Deadline_Miss"}).reset_index()

    # sort for rolling mean
    if not summary_df.empty and "episode" in summary_df.columns:
        summary_df = summary_df.sort_values("episode").reset_index(drop=True)

    if not summary_df.empty:
        total = (summary_df["OnTime"] + summary_df["Deadline_Miss"]).replace(0, 1)

        # rates
        summary_df["OnTime_Rate"] = summary_df["OnTime"] / total
        summary_df["Deadline_Miss_Rate"] = summary_df["Deadline_Miss"] / total

        # rolling averages (window=40, safe for fewer episodes)
        summary_df["AVG_OnTime_Rate"] = summary_df["OnTime_Rate"].rolling(window=40, min_periods=1).mean()
        summary_df["AVG_Deadline_Miss_Rate"] = summary_df["Deadline_Miss_Rate"].rolling(window=40, min_periods=1).mean()
    else:
        summary_df["OnTime_Rate"] = []
        summary_df["Deadline_Miss_Rate"] = []
        summary_df["AVG_OnTime_Rate"] = []
        summary_df["AVG_Deadline_Miss_Rate"] = []

    # ---------------------------
    # Write Excel
    # ---------------------------
    with pd.ExcelWriter(filename) as writer:
        df_params.to_excel(writer, sheet_name="Params", index=False)
        task_df.to_excel(writer, sheet_name="Tasks", index=False)
        server_info.to_excel(writer, sheet_name="Servers", index=False)
        df_logs.to_excel(writer, sheet_name="Logs", index=False)
        df_task_Assignments.to_excel(writer, sheet_name="TaskAssignments", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # ---------------------------
    # Add Excel-native charts
    # ---------------------------
    wb = load_workbook(filename)

    # Summary: line charts for AVG_OnTime_Rate and AVG_Deadline_Miss_Rate
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        max_row = summary_df.shape[0] + 1  # header included

        # columns:
        # 1 episode | 2 OnTime | 3 Deadline_Miss | 4 OnTime_Rate | 5 Deadline_Miss_Rate
        # 6 AVG_OnTime_Rate | 7 AVG_Deadline_Miss_Rate
        if max_row >= 2:
            # Chart 1: AVG_OnTime_Rate
            line1 = LineChart()
            line1.title = "Average On-Time Rate Over Episodes (Rolling 40)"
            line1.x_axis.title = "Episode"
            line1.y_axis.title = "AVG_OnTime_Rate"

            data1 = Reference(ws_sum, min_col=6, min_row=1, max_col=6, max_row=max_row)  # AVG_OnTime_Rate
            cats = Reference(ws_sum, min_col=1, min_row=2, max_row=max_row)              # episode

            line1.add_data(data1, titles_from_data=True)
            line1.set_categories(cats)
            line1.width = 22
            line1.height = 10
            ws_sum.add_chart(line1, "I2")

            # Chart 2: AVG_Deadline_Miss_Rate
            line2 = LineChart()
            line2.title = "Average Deadline-Miss Rate Over Episodes (Rolling 40)"
            line2.x_axis.title = "Episode"
            line2.y_axis.title = "AVG_Deadline_Miss_Rate"

            data2 = Reference(ws_sum, min_col=7, min_row=1, max_col=7, max_row=max_row)  # AVG_Deadline_Miss_Rate
            line2.add_data(data2, titles_from_data=True)
            line2.set_categories(cats)
            line2.width = 22
            line2.height = 10
            ws_sum.add_chart(line2, "I20")

    # Logs charts (Rewards + optional Delay)
    if "Logs" in wb.sheetnames:
        ws_logs = wb["Logs"]
        header = [cell.value for cell in ws_logs[1]]

        def col_idx(name):
            try:
                return header.index(name) + 1
            except ValueError:
                return None

        c_episode = col_idx("Episode")
        c_avg_reward = col_idx("Avg Reward")
        c_ep_reward = col_idx("Episode Reward")
        c_avg_delay = col_idx("Avg Delay")

        max_row_logs = ws_logs.max_row

        # Rewards chart
        if c_episode and (c_avg_reward or c_ep_reward) and max_row_logs >= 2:
            rewards_chart = LineChart()
            rewards_chart.title = "Rewards per Episode"
            rewards_chart.y_axis.title = "Reward"
            rewards_chart.x_axis.title = "Episode"

            min_col = min([c for c in [c_avg_reward, c_ep_reward] if c is not None])
            max_col = max([c for c in [c_avg_reward, c_ep_reward] if c is not None])

            data = Reference(ws_logs, min_col=min_col, min_row=1, max_col=max_col, max_row=max_row_logs)
            cats = Reference(ws_logs, min_col=c_episode, min_row=2, max_row=max_row_logs)

            rewards_chart.add_data(data, titles_from_data=True)
            rewards_chart.set_categories(cats)

            ws_logs.add_chart(rewards_chart, "F2")

        # Delay chart
        if c_episode and c_avg_delay and max_row_logs >= 2:
            delay_chart = LineChart()
            delay_chart.title = "Avg Delay per Episode"
            delay_chart.y_axis.title = "Delay"
            delay_chart.x_axis.title = "Episode"

            data = Reference(ws_logs, min_col=c_avg_delay, min_row=1, max_col=c_avg_delay, max_row=max_row_logs)
            cats = Reference(ws_logs, min_col=c_episode, min_row=2, max_row=max_row_logs)

            delay_chart.add_data(data, titles_from_data=True)
            delay_chart.set_categories(cats)

            ws_logs.add_chart(delay_chart, "F20")

    wb.save(filename)
    print("successfully saved logs !")
